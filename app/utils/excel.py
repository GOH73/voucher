import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_voucher_transfer(file_path, data):
    """
    创建凭证移交清单的Excel文件
    :param file_path: Excel文件保存路径
    :param data: 凭证数据，包含90行，每行(序号, 月份, 凭证号, 遗失)
    """
    # 创建一个新的工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "凭证移交清单"

    # 设置列宽
    column_widths = [
        6, 6, 10, 6, 3,
        6, 6, 10, 6, 3,
        6, 6, 10, 6
    ]  # 根据需要调整
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 定义表头
    headers = [
        "序号", "月份", "凭证号", "遗失", "",
        "序号", "月份", "凭证号", "遗失", "",
        "序号", "月份", "凭证号", "遗失"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # 设置表头字体为加粗
        cell.font = Font(bold=True)
        # 居中对齐
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 添加边框
        if col_num % 5 == 0:
            continue
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # 填充数据
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            # 居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 添加边框
            if col_num % 5 == 0:
                continue
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # 保存Excel文件
    wb.save(file_path)
    print(f"Excel文件已生成：{file_path}")


def generate_sample_data():
    """
    生成示例数据，实际应用中应从数据库中提取
    :return: 包含90行数据的列表
    """
    data = []
    for i in range(1, 31):
        data.append((i, "", '', '', '', 30 + i, "", '', '', '', 60 + i, "", '', ''))
    # 修正月份逻辑，确保每列30行
    # 由于需要三列，每列30行，这里简化处理，所有月份为空或根据需要填充
    # 实际应用中应根据具体逻辑填充月份
    return data


if __name__ == "__main__":
    # 生成示例数据
    data = generate_sample_data()

    # 创建Excel文件
    create_excel_voucher_transfer("凭证移交清单.xlsx", data)
