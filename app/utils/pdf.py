from reportlab.pdfbase import pdfmetrics  # 注册字体
from reportlab.pdfbase.ttfonts import TTFont  # 字体类
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
from reportlab.platypus import Table

pdfmetrics.registerFont(TTFont('simkai', 'simkai.ttf'))


def excel_to_pdf_manual(excel_path, pdf_path):
    # 读取 Excel 数据
    wb = load_workbook(excel_path)
    ws = wb.active

    # 提取数据（假设数据在第一个 Sheet）
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # 创建 PDF
    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter

    # 添加表头（示例）
    c.drawString(100, height - 50, "凭证移交清单")

    # 添加表格（简化版，实际需要调整位置）
    table_data = [["序号", "月份", "凭证号", "遗失"]] + [list(row) for row in data if row[0] is not None]
    table = Table(table_data)
    # 设置表格样式（略）

    # 保存 PDF
    c.save()


# 使用示例
excel_to_pdf_manual("凭证移交清单.xlsx", "凭证移交清单.pdf")
